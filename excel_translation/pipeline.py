"""
excel_translation.pipeline

Implements the translation pipeline core:
extract translatable cells -> translate (with token protection, chunking, cache) -> map results per cell.

Excel read/write (TASK-014 onward) may depend on extra Excel libraries.
"""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
import os
import time
from typing import Any, Dict, Iterable, Optional, Tuple

from .cache.store import TranslationCache
from .chunking import make_batches
from .glossary.loader import load_glossary
from .models.contracts import (
    CellKey,
    TranslationReport,
    TranslationResult,
    TranslationStatus,
    TranslationRunMeta,
)
from .reader import extract_translatable_cells
from .token_protection import protect_tokens, restore_tokens
from .error_policy import default_policy, build_failed_result
from .translator.base import TranslatorBackend, TranslateBatchRequest
from .translator.mock import MockTranslator
from .writer import write_translations
from .output_naming import resolve_output_path


@dataclass(frozen=True)
class TranslationRunConfig:
    input_path: str
    output_path: str
    source_lang: str = "ja"
    target_lang: str = "vi"
    glossary_path: str | None = None
    token_policy_mode: str = "default"
    formula_mode: str = "skip"
    cache_enabled: bool = True
    # error_policy_mode: "continue-on-error" | "fail-fast"
    error_policy_mode: str = "continue-on-error"
    cache_ttl_seconds: int | None = None


def _compute_cache_key(
    source_text: str,
    source_lang: str,
    target_lang: str,
    glossary_version: str,
    token_policy_id: str,
) -> str:
    raw = source_text + source_lang + target_lang + glossary_version + token_policy_id
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _load_glossary_if_needed(glossary_path: str | None) -> Tuple[dict[str, str], str]:
    if not glossary_path:
        return {}, "none"
    res = load_glossary(glossary_path)
    return res.glossary, res.glossary_version


def translate_workbook(
    *,
    workbook: Any,
    config: TranslationRunConfig,
    translator: TranslatorBackend | None = None,
    cache: TranslationCache | None = None,
) -> Tuple[Dict[CellKey, TranslationResult], TranslationReport]:
    """
    Run extraction + translation + mapping on an already loaded workbook object.
    """

    start_ts = time.time()
    translator = translator or MockTranslator()
    glossary_data, glossary_version = _load_glossary_if_needed(config.glossary_path)
    cache = cache or (TranslationCache(path=None, ttl_seconds=config.cache_ttl_seconds) if config.cache_enabled else TranslationCache(path=None, ttl_seconds=None))

    candidates = extract_translatable_cells(workbook, formula_mode=config.formula_mode)
    translatable_cells = len(candidates)

    # Note: total_string_cells is an approximate metric at this stage.
    # TASK-016 will refine reporting/logging.
    total_string_cells = translatable_cells

    # Deduplicate by original text (selection rule uses "string + Japanese heuristic" only).
    unique_texts: list[str] = []
    seen: set[str] = set()
    for c in candidates:
        if c.original_text not in seen:
            seen.add(c.original_text)
            unique_texts.append(c.original_text)

    token_policy_id = config.token_policy_mode

    # Cache check (per unique original text).
    cache_keys = [
        _compute_cache_key(
            source_text=t,
            source_lang=config.source_lang,
            target_lang=config.target_lang,
            glossary_version=glossary_version,
            token_policy_id=token_policy_id,
        )
        for t in unique_texts
    ]

    cached_hits: dict[str, str] = {}
    if config.cache_enabled:
        cache_get = cache.get_many(cache_keys)
        cached_hits = cache_get.hits

    policy = default_policy(config.error_policy_mode)

    # Build text -> translated_text map
    translated_by_text: dict[str, str] = {}
    translated_success_texts: list[str] = []
    failed_error_by_text: dict[str, str] = {}

    # Translate misses
    texts_to_translate: list[str] = []
    cache_key_by_text: dict[str, str] = {}
    for text, key in zip(unique_texts, cache_keys):
        cache_key_by_text[text] = key
        if config.cache_enabled and key in cached_hits:
            translated_by_text[text] = cached_hits[key]
        else:
            texts_to_translate.append(text)

    # Chunk translation requests.
    max_chars_per_batch = 3000  # best-effort default; can be made configurable later
    batches = make_batches(texts_to_translate, max_chars=max_chars_per_batch)

    # Translate each batch, but isolate failures per text.
    for batch in batches:
        protected_texts: list[str] = []
        protection_maps: list[dict[str, str]] = []
        for text in batch:
            protected = protect_tokens(text)
            protected_texts.append(protected.protected_text)
            protection_maps.append(protected.mapping)

        req = TranslateBatchRequest(
            texts=protected_texts,
            source_lang=config.source_lang,
            target_lang=config.target_lang,
            glossary=glossary_data if glossary_data else None,
            token_policy_mode=token_policy_id,
        )

        try:
            resp = translator.translate_batch(req)
            protected_translations = resp.translations
            if len(protected_translations) != len(batch):
                raise RuntimeError("Translator returned mismatched translations count.")

            for orig_text, prot_translation, prot_map in zip(batch, protected_translations, protection_maps):
                translated = restore_tokens(prot_translation, prot_map)
                translated_by_text[orig_text] = translated
                translated_success_texts.append(orig_text)
        except Exception:
            # Fallback to per-text calls to support per-cell fallback policy.
            for text, prot_text, prot_map in zip(batch, protected_texts, protection_maps):
                try:
                    single_req = TranslateBatchRequest(
                        texts=[prot_text],
                        source_lang=config.source_lang,
                        target_lang=config.target_lang,
                        glossary=glossary_data if glossary_data else None,
                        token_policy_mode=token_policy_id,
                    )
                    single_resp = translator.translate_batch(single_req)
                    if len(single_resp.translations) != 1:
                        raise RuntimeError("Translator returned mismatched translations count.")
                    restored = restore_tokens(single_resp.translations[0], prot_map)
                    translated_by_text[text] = restored
                    translated_success_texts.append(text)
                except Exception as exc:
                    failed_error_by_text[text] = str(exc)
                    if policy.mode == "fail-fast":
                        raise

    # Cache successful translations only.
    if config.cache_enabled:
        to_cache: dict[str, str] = {}
        for text in set(translated_success_texts):
            key = cache_key_by_text[text]
            to_cache[key] = translated_by_text[text]
        cache.set_many(to_cache)

    # Map results per cell
    translation_mapping: Dict[CellKey, TranslationResult] = {}
    translated_success_cells = 0
    translated_failed_cells = 0
    failures: list[TranslationResult] = []

    for c in candidates:
        translated_text = translated_by_text.get(c.original_text)
        if translated_text is not None:
            status = TranslationStatus.translated_success
            translated_success_cells += 1
            translation_mapping[c.cell_key] = TranslationResult(
                cell_key=c.cell_key,
                original_text=c.original_text,
                translated_text=translated_text,
                status=status,
                reason=None,
                error_message=None,
            )
        else:
            status = TranslationStatus.translated_failed
            translated_failed_cells += 1
            failure = build_failed_result(
                cell_key=c.cell_key,
                original_text=c.original_text,
                error_message=failed_error_by_text.get(c.original_text),
            )
            translation_mapping[c.cell_key] = failure
            failures.append(failure)

    run_id = str(int(start_ts * 1000))
    meta: TranslationRunMeta = {
        "input_filename": os.path.basename(config.input_path) if config.input_path else "unknown",
        "output_filename": os.path.basename(config.output_path) if config.output_path else "unknown",
        "runtime_ms": int((time.time() - start_ts) * 1000),
        "glossary_version": glossary_version,
        "token_policy_id": token_policy_id,
        "formula_mode": config.formula_mode,
    }

    report = TranslationReport(
        run_id=run_id,
        total_string_cells=total_string_cells,
        translatable_cells=translatable_cells,
        translated_success=translated_success_cells,
        translated_failed=translated_failed_cells,
        skipped=max(0, total_string_cells - translatable_cells),
        failures=failures,
        meta=dict(meta),
    )

    return translation_mapping, report


def translate_excel(config: TranslationRunConfig) -> Dict[str, Any]:
    """
    File-based entrypoint.

    This function requires an Excel loader library (e.g., `openpyxl`).
    """

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "Excel loading requires `openpyxl`, but it is not available in this environment. "
            "Call `translate_workbook()` with a preloaded workbook object instead."
        ) from exc

    # We must preserve formulas in the output file (MVP requirement).
    # For `cached-result-only`, we read cached values via `data_only=True` *only for extraction*,
    # but we always write into a workbook loaded with `data_only=False` to keep formulas intact.
    workbook_for_write = load_workbook(config.input_path, data_only=False)

    workbook_for_read = workbook_for_write
    if config.formula_mode == "cached-result-only":
        workbook_for_read = load_workbook(config.input_path, data_only=True)

    translation_mapping, report = translate_workbook(workbook=workbook_for_read, config=config)

    output_path = resolve_output_path(
        input_path=config.input_path,
        output_path=config.output_path,
        target_lang=config.target_lang,
        overwrite=False,
    )

    out_workbook = write_translations(workbook_for_write, translation_mapping)

    if not hasattr(out_workbook, "save"):
        raise RuntimeError("Excel workbook object does not support .save(); cannot export output file.")
    out_workbook.save(output_path)

    return {"report": report, "translation_mapping": translation_mapping, "output_path": output_path}

